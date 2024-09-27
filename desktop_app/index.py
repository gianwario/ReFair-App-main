from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
import os
import openpyxl
import json
from PIL import Image, ImageTk

from components.user_story_canvas import UserStoryCanvas
from components.custom_messagebox import CustomMessageBox 
from colors import COLORS 
from domain_utils import getDomain, getMLTask, feature_extraction

# Variabili globali
user_story_listbox = None
scrollbar = None
file_path = ""
user_stories = []
canvas_frame = None
scrollbar = None

class RefairDesktopApp:
    def __init__(self, root):
        self.root = root
        self.root.title("ReFair desktop app")
        self.root.geometry("1100x600")  # Dimensioni della finestra
        self.root.iconbitmap('desktop_app/icons/bill_invoice_shop_icon.ico')

        # Path all'icona (modifica il percorso se necessario)
        icon_right_arrow_path = "desktop_app/icons/right_arrow_icon.png"
        document_attach_outline_path = "desktop_app/icons/document-attach-outline.png"
        cloud_download_outline_path = "desktop_app/icons/code-download-outline.png"
        cloud_upload_outline_path = "desktop_app/icons/cloud-upload-outline.png"
        analytics_outline_path = "desktop_app/icons/analytics-outline.png"

        # Carica l'icona
        right_arrow_icon = Image.open(icon_right_arrow_path)
        right_arrow_icon = right_arrow_icon.resize((20, 20))
        self.right_arrow = ImageTk.PhotoImage(right_arrow_icon)

        document_attach_outline_icon = Image.open(document_attach_outline_path)
        document_attach_outline_icon = document_attach_outline_icon.resize((20, 20))
        self.document_attach_outline = ImageTk.PhotoImage(document_attach_outline_icon)

        cloud_upload_outline_icon = Image.open(cloud_upload_outline_path)
        cloud_upload_outline_icon = cloud_upload_outline_icon.resize((20, 20))
        self.cloud_upload_outline = ImageTk.PhotoImage(cloud_upload_outline_icon)

        cloud_download_outline_icon = Image.open(cloud_download_outline_path)
        cloud_download_outline_icon = cloud_download_outline_icon.resize((20, 20))
        self.cloud_download_outline = ImageTk.PhotoImage(cloud_download_outline_icon)

        analytics_outline_icon = Image.open(analytics_outline_path)
        analytics_outline_icon = analytics_outline_icon.resize((20, 20))
        self.analytics_outline = ImageTk.PhotoImage(analytics_outline_icon)

        # Crea uno stile personalizzato per il bottone
        style = ttk.Style()
        style.configure("Custom.TButton",
                        bordercolor="black",
                        borderwidth=2,
                        relief="solid",
                        padding=5,
                        background='black') 

        # Sidebar
        sidebar = Frame(root, width=200, background=COLORS['light_gray'])
        sidebar.pack(fill=Y, side=LEFT)

        # Frame contenitore principale per i contenuti
        #content_frame = Frame(root)
        #content_frame.pack(expand=True, fill=BOTH, side=LEFT)

        # Main content area
        self.main_content = Frame(root, background='white')
        self.main_content.pack(fill=BOTH, expand=True)

        ##SubFrames di Main Content
        select_frame = Frame(self.main_content, background=COLORS['background'])
        select_frame.grid(row=0, column=0)

        entry_frame = Frame(self.main_content, background=COLORS['background'])
        entry_frame.grid(row=1, column=0)

        # Frame per refair_info
        refair_info = Frame(root, background=COLORS['background'])

        # Frame per refair_suggestions
        refair_suggestions = Frame(root, background=COLORS['background'])


        #Sidebar content
        main_button = Button(sidebar, text="ReFair App", image=self.right_arrow, compound=LEFT, command=self.show_main_content,
                            borderwidth=0, highlightthickness=0, relief=FLAT, padx=10, anchor='w', background=COLORS['light_gray'],
                                activebackground=COLORS['light_gray'], activeforeground=COLORS['background'])
        main_button.pack(fill=X, pady=20)

        info_button = Button(sidebar, text="ReFair in a nutshell", image=self.right_arrow, compound=LEFT, command=self.show_refair_info, 
                            borderwidth=0, highlightthickness=0, relief=FLAT, padx=10, anchor='w', background=COLORS['light_gray'],
                                activebackground=COLORS['light_gray'], activeforeground=COLORS['background'])
        info_button.pack(fill=X, pady=20)

        suggestions_button = Button(sidebar, text="How to use ReFair", image=self.right_arrow, compound=LEFT, command=self.show_refair_suggestions,
                                    borderwidth=0, highlightthickness=0, relief=FLAT, padx=10, anchor='w', background=COLORS['light_gray'],
                                activebackground=COLORS['light_gray'], activeforeground=COLORS['background'])
        suggestions_button.pack(fill=X, pady=20)


        #Main_content frame 
        ##Select_frame content
        select_file_button = ttk.Button(select_frame, text="Select file", image=self.document_attach_outline, compound=LEFT, style="Custom.TButton", command=self.open_file)
        select_file_button.grid(row=0, column=0, padx=10, pady=20)

        self.file_name_label = Label(select_frame, text="", background=COLORS['background'])
        self.file_name_label.grid(row=0, column=1, padx=10, pady=20)

        space_label = Label(select_frame, text="", background=COLORS['background'])
        space_label.grid(row=0, column=2, padx=100)

        load_button = ttk.Button(select_frame, text="Load", image=self.cloud_upload_outline, compound=LEFT, style="Custom.TButton", command=self.load_file)
        load_button.grid(row=0, column=3, padx=10, pady=10)

        save_button = ttk.Button(select_frame, text="Download all", image=self.cloud_download_outline, compound=LEFT, style="Custom.TButton", command=self.save_as_json)
        save_button.grid(row=0, column=4, padx=10, pady=10)

        ##Entry_frame content
        # Entry
        self.entry_us = ttk.Entry(entry_frame, width=100)
        self.entry_us.grid(row=0, column=0, padx=20)

        self.analyze_button = ttk.Button(entry_frame, text="Analyze", image=self.analytics_outline, compound=LEFT, command=lambda: self.analyze(self.entry_us.get()))
        self.analyze_button.grid(row=0, column=1)

        # Inizialmente disabilita il pulsante
        self.analyze_button.config(state=DISABLED)

        # Configura il controllo dell'entry
        self.entry_us_var = StringVar()
        self.entry_us.config(textvariable=self.entry_us_var)
        self.entry_us_var.trace_add('write', self.validate_entry)

        ##USs title
        us_title_label = Label (self.main_content, text="User Stories", background=COLORS['background'], font=('Helvetica', 16))
        us_title_label.grid(row=2, column=0)

        # Titolo del frame
        info_frame_title = Label(refair_info, background=COLORS['background'], text="ReFair in a nutshell", font=('Helvetica', 18))
        info_frame_title.pack(pady=10)

        # Sezione "What is ReFair?"
        what_is_refair_title = Label(refair_info, background=COLORS['background'], text="What is ReFair?", font=('Helvetica', 16))
        what_is_refair_title.pack(pady=20)

        what_is_refair_content = Label(refair_info, background=COLORS['background'], wraplength=900, font=('Helvetica', 10), justify="left", text="ReFair is an innovative context-aware automated framework designed to support fairness requirements engineering. It utilizes natural language processing (NLP) and word embedding techniques to identify sensitive features in user stories (USs), alerting developers early on to potential concerns.")
        what_is_refair_content.pack(anchor="w", padx=10)

        # Sezione "Main functionalities"
        main_functionalities_title = Label(refair_info, background=COLORS['background'], text="Main functionalities", font=('Helvetica', 16))
        main_functionalities_title.pack(pady=20)

        main_functionalities_text = """ReFair is a model that consists of two main components:\n
        - Application Domain Classification. This component is responsible for classifying the most likely application domain of the US among the 34 domains available in the ontology; \n
        - Machine Learning Tasks Classification. This is responsible for classifying the ML tasks likely to be employed when implementing the US. The problem has been modeled as a multi-label classification task, as a US may be operationalized using multiple ML techniques; \n
        - Sensitive Features Recommendation. The application domain and ML tasks classified in the previous step are finally used to recommend sensitive features. ReFair exploits the base ontology to identify the sensitive features connected to both the application domain and ML tasks concerned with the classified domain. The intersection of those sensitive features represents the final outcome of the framework. In other terms the outcome comprises the set of sensitive features relevant when jointly considering the application domain and the learning tasks."""

        main_functionalities_content = Label(refair_info, background=COLORS['background'], wraplength=900, font=('Helvetica', 10), justify="left", text=main_functionalities_text)
        main_functionalities_content.pack(anchor="w", padx=10)

        # Sezione "Technical aspects"
        technical_aspects_title = Label(refair_info, background=COLORS['background'], text="Technical aspects", font=('Helvetica', 16))
        technical_aspects_title.pack(pady=20)

        technical_aspects_content = Label(refair_info, background=COLORS['background'], wraplength=900, font=('Helvetica', 10), justify="left", text="The framework has been designed to be conservative enough and identify all the potential ML tasks that may lead to unfairness. From a practical perspective, this choice may allow the users to receive a larger set of sensitive features, hence favoring recall over precision.")
        technical_aspects_content.pack(anchor="w", padx=10)

        # Contenuto del frame refair_suggestions
        usage_frame_title = Label(refair_suggestions, background=COLORS['background'], text="How to use ReFair", font=('Helvetica', 18))
        usage_frame_title.pack(pady=10)

        self.refair_info = refair_info
        self.refair_suggestions = refair_suggestions

        recommendation_title = Label(refair_suggestions, background=COLORS['background'], text="Recommendation", font=('Helvetica', 16))
        recommendation_title.pack(pady=20)

        recommendation_text = """To properly run the ReFair analysis, you should upload a file that meets specific conditions:\n
        - The file should be in xlsx format;\n
        - The spreadsheet can contain an arbitrary number of columns, but at least one column should be named "User Story" and should contain all the User Stories you want to be analyzed."""

        recommendation_content = Label(refair_suggestions, background=COLORS['background'], wraplength=900, font=('Helvetica', 10), justify="left", text=recommendation_text)
        recommendation_content.pack(anchor="w", padx=10)


        functionalities_title = Label(refair_suggestions, background=COLORS['background'], text="Web-app functionalities", font=('Helvetica', 16))
        functionalities_title.pack(pady=20)

        functionalities_text = """In detail:\n
        - The Select Button allows you to select a User Stories spreadsheet from your machine;\n
        - The Load Button allows you to upload the User Stories spreadsheet;\n
        - The Download all Button allows you to download a structured JSON report containing the results for all the User Stories analyzed by ReFair;\n
        - The Analyze Button allows you to visualize the ReFair analysis for a single User Story;\n
        - The Download Button (in the pop-up window) allows you to download a structured JSON report containing the results of the single User Story analyzed by ReFair."""

        functionalities_content = Label(refair_suggestions, background=COLORS['background'], wraplength=900, font=('Helvetica', 10), justify="left", text=functionalities_text)
        functionalities_content.pack(anchor="w", padx=10)

        # Mostra il main_content all'avvio
        self.show_frame(self.main_content)

    def open_file(self):
        global file_path
        #Needed to delete the USs of the previous file that the model has to analyze when selecting a new file
        user_stories.clear()

        # Apri la finestra di dialogo per selezionare un file
        file_path = filedialog.askopenfilename()
        
        # Controlla se il file selezionato è un .xlsx
        if file_path.endswith('.xlsx'):
            # Se è un file .xlsx, crea una label con il nome del file
            file_name = os.path.basename(file_path)
            self.file_name_label.config(text=f"{file_name}", fg="black")
        else:
            # Se non è un file .xlsx, mostra un messaggio di errore
            messagebox.showerror(title="Error", message="The file selected isn't a .xlsx file")
            file_path = ""
            user_stories.clear()  # Cancella le user stories salvate

    def load_file(self):
        global user_stories, canvas_frame

        if not file_path:
            # Se non è stato selezionato alcun file valido, mostra un messaggio di errore
            messagebox.showerror(title="Error", message="No .xlsx file selected")
            return
        
        # Carica il file .xlsx
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        # Trova la colonna "User Story"
        user_story_col = None
        for col in sheet.iter_cols(1, sheet.max_column):
            if col[0].value == "User Story":
                user_story_col = col
                break
        
        if user_story_col is None:
            # Se la colonna "User Story" non esiste, mostra un messaggio di errore
            messagebox.showerror(title="Error", message="The 'User Story' column has not been found")
            return

        # Cancella eventuali user stories precedenti
        user_stories.clear()

        # Rimuovi il Frame precedente con il Canvas, se esiste
        if canvas_frame is not None:
            canvas_frame.destroy()

        # Aggiungi tutte le User Stories alla lista
        for cell in user_story_col[1:]:  # Salta l'intestazione
            user_stories.append(cell.value)

        # Crea e posiziona la nuova lista custom UserStoryCanvas. 
        # Ciò è necessario in quanto ogni elemento della lista è a sua volta composto da frame
        canvas_frame = UserStoryCanvas(self.main_content, user_stories)
        canvas_frame.grid(row=3, column=0, padx=10, pady=10, sticky='nsew')

        # Configura le colonne e righe della griglia per l'espansione
        self.main_content.grid_columnconfigure(0, weight=1)
        self.main_content.grid_columnconfigure(1, weight=0)
        self.main_content.grid_rowconfigure(3, weight=1)

    def save_as_json(self):
        if not user_stories:
            # Se non ci sono user stories caricate, mostra un messaggio di errore
            messagebox.showerror(title="Error", message="There are no User Stories to download")
            return

        # Assumiamo che user_stories sia una lista di dizionari
        all_data = []

        for user_story in user_stories:
            predicted_domain = getDomain(user_story)
            predicted_task = getMLTask(user_story, predicted_domain)
            results = feature_extraction(predicted_domain, predicted_task)
            
            # Organizza i dati per ogni User Story
            data = {
                "user_story": user_story,
                "story_domain": predicted_domain,
                "sensitive_features": results
            }
            all_data.append(data)

        # Permette di salvare il file JSON con un nome specifico
        file_path = filedialog.asksaveasfilename(defaultextension=".json",
                                                filetypes=[("JSON files", "*.json"),
                                                        ("All files", "*.*")])
        
        if file_path:
            # Salva tutti i dati in un file JSON
            with open(file_path, 'w') as json_file:
                json.dump(all_data, json_file, indent=4)
            messagebox.showinfo("Success", "Results have been downloaded successfully!")


    def analyze(self, user_story):
        # Chiama la funzione getDomain e aggiorna la label con il risultato
        predicted_domain = getDomain(user_story)
        predicted_task = getMLTask(user_story, predicted_domain)
        results = feature_extraction(predicted_domain, predicted_task)
        # Richiamo il custom component strutturato come una messagebox a cui passo tutti i dati necessari
        CustomMessageBox(root, predicted_domain, user_story, predicted_domain, results)

    def validate_entry(self, *args):
        entry_content = self.entry_us.get()
        if entry_content.strip():  # Se non è vuoto o solo spazi. Serve per non far partire l'analyze all'avvio della desktop app
            self.analyze_button.config(state=NORMAL)
        else:
            self.analyze_button.config(state=DISABLED)

    def show_frame(self, frame):
        frame.tkraise()

    # Funzioni che saranno chiamate quando si clicca sui pulsanti della sidebar
    def show_main_content(self):
        self.show_frame(self.main_content)
        self.main_content.pack(fill=BOTH, expand=True)
        self.refair_info.pack_forget()
        self.refair_suggestions.pack_forget()
        

    def show_refair_info(self):
        self.show_frame(self.refair_info)
        self.refair_info.pack(fill=BOTH, expand=True)
        self.main_content.pack_forget()
        self.refair_suggestions.pack_forget()

    def show_refair_suggestions(self):
        self.show_frame(self.refair_suggestions)
        self.refair_suggestions.pack(fill=BOTH, expand=True)
        self.refair_info.pack_forget()
        self.main_content.pack_forget()


if __name__ == '__main__':
    root = Tk() # Finestra principale
    app = RefairDesktopApp(root)
    root.mainloop() # Necessario per rendere la finestra visibile
